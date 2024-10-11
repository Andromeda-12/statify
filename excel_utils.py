import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from loguru import logger


def create_backup(file_name):
    """Создает резервную копию файла с указанием даты"""
    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file_name = f"{file_name}_backup_{current_date}.xlsx"
    shutil.copy(file_name, backup_file_name)
    logger.info(f"Создана резервная копия файла: {backup_file_name}")


def create_or_update_excel_report(establishments_data, search_rankings_by_date):
    """
    Создает или обновляет Excel-отчет по заведениям в транспонированном виде.
    Структура: первая строка "Запрос", затем идут запросы; вторая строка "Частота", затем частоты запросов.
    Далее идут строки с датами, а в колонках — позиции запросов.

    Если функция вызывается несколько раз, данные обновляются и суммируются.

    establishments_data: Список заведений и запросов
    search_rankings_by_date: Данные о позициях и частоте запросов за каждый день
    """
    # Определяем текущий месяц и год
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    days_in_month = (
        (datetime(current_year, current_month + 1, 1) - timedelta(days=1)).day
        if current_month != 12
        else 31
    )

    # Имя файла для текущего месяца
    file_name = f"Отчет_по_заведениям_{current_month:02}_{current_year}.xlsx"

    if os.path.exists(file_name):
        create_backup(file_name)

    # Если файл не существует, создаем новую таблицу
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Удаляем автоматически созданный лист
        logger.info(f"Создан новый Excel файл '{file_name}' для текущего месяца.")
    else:
        wb = openpyxl.load_workbook(file_name)

    # Проходим по каждому заведению
    for establishment in establishments_data:
        establishment_id = establishment["id"]
        queries = establishment["queries"]

        # Проверяем, существует ли лист для заведения, если нет — создаем
        if establishment_id not in wb.sheetnames:
            ws = wb.create_sheet(title=establishment_id)

            # Устанавливаем первую строку: заголовок "Запрос", а затем запросы
            headers = ["Запрос"] + [query for query in queries]
            ws.append(headers)

            # Устанавливаем вторую строку: заголовок "Частота" и частоты запросов
            frequency_row = ["Частота"] + [0] * len(queries)
            ws.append(frequency_row)

            # Устанавливаем ширину колонок
            ws.column_dimensions["A"].width = 20  # Столбец "Запрос"/"Частота"
            for col in range(2, len(headers) + 1):  # Колонки с запросами
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 20
        else:
            ws = wb[establishment_id]

        # Создаем словарь для поиска колонок запросов
        existing_queries = {
            ws.cell(row=1, column=col).value: col for col in range(2, ws.max_column + 1)
        }

        # Создаем словарь для поиска существующих строк по дате
        existing_dates = {
            ws.cell(row=row, column=1).value: row for row in range(3, ws.max_row + 1)
        }

        # Добавляем или обновляем данные по датам
        for day in range(1, days_in_month + 1):
            date_str = f"{day:02}.{current_month:02}.{current_year}"

            if date_str in existing_dates:
                # Если дата уже существует, обновляем значения
                row_index = existing_dates[date_str]
            else:
                # Если дата не существует, создаем новую строку
                row_index = ws.max_row + 1
                ws.append([date_str] + [""] * len(queries))
                existing_dates[date_str] = row_index

            # Обновляем позиции для каждого запроса
            if date_str in search_rankings_by_date:
                for query in queries:
                    col_index = existing_queries.get(query)

                    if (
                        establishment_id in search_rankings_by_date[date_str]
                        and query in search_rankings_by_date[date_str][establishment_id]
                    ):
                        position_data = search_rankings_by_date[date_str][
                            establishment_id
                        ][query]

                        # Извлекаем частоту и позицию
                        frequency = position_data["frequency"]
                        position = position_data["position"]
                        
                        # Обновляем частоту во второй строке
                        current_frequency = ws.cell(row=2, column=col_index).value
                        ws.cell(row=2, column=col_index).value = (
                            current_frequency + frequency
                        )

                        # Обновляем ячейку с позицией запроса для данной даты
                        current_position = ws.cell(
                            row=row_index, column=col_index
                        ).value

                        # Если ячейка уже содержит значение, суммируем его с новым
                        if current_position not in (None, "", 0):
                            ws.cell(row=row_index, column=col_index).value = (
                                current_position
                            )
                        else:
                            ws.cell(row=row_index, column=col_index).value = position

    # Сохраняем изменения в файле
    wb.save(file_name)
    logger.info(f"Excel файл '{file_name}' успешно обновлен.")

    return file_name
